from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import os
from datetime import datetime
import json
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Register Telr webhook blueprint
try:
    from telr_webhook import telr_webhook_bp
    app.register_blueprint(telr_webhook_bp)
    print("✅ Telr webhook blueprint registered")
except ImportError:
    print("⚠️ Telr webhook module not found - payment webhooks will not work")

# Register Telr API proxy blueprint
try:
    from telr_api import telr_api_bp
    app.register_blueprint(telr_api_bp)
    print("✅ Telr API proxy blueprint registered")
except ImportError:
    print("⚠️ Telr API module not found - payment API will not work")

# Hotel configuration
HOTEL_EXCEL_FILE_PATH = 'hotels.xlsx'
HOTEL_SHEET_NAME = 'Hotels'
HOTEL_HEADERS = ['Hotel Code', 'Name', 'Rating', 'Address', 'City ID', 'Country Code', 
                 'Latitude', 'Longitude', 'Facilities', 'Images', 'Created At']

# Room configuration
ROOM_EXCEL_FILE_PATH = 'hotel_rooms.xlsx'
ROOM_SHEET_NAME = 'Rooms'
ROOM_HEADERS = ['Room ID', 'Hotel Code', 'Booking Code', 'Room Name', 'Base Price', 
                'Total Fare', 'Currency', 'Is Refundable', 'Day Rates', 'Extras', 'Created At']

# Wishlist configuration
WISHLIST_EXCEL_FILE_PATH = 'wishlist.xlsx'
WISHLIST_SHEET_NAME = 'Wishlist'
WISHLIST_HEADERS = ['Wishlist ID', 'Customer ID', 'Hotel Code', 'Hotel Name', 'Hotel Rating', 
                    'Address', 'City', 'Country', 'Price', 'Currency', 'Image URL', 'Search Params', 'Created At']

def create_hotel_excel_file_if_not_exists():
    """Create hotel Excel file with headers if it doesn't exist"""
    if not os.path.exists(HOTEL_EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = HOTEL_SHEET_NAME
        
        # Add headers
        for col, header in enumerate(HOTEL_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Save the workbook
        wb.save(HOTEL_EXCEL_FILE_PATH)
        print(f"Created new hotel Excel file: {HOTEL_EXCEL_FILE_PATH}")

def save_hotel_to_excel(data):
    """Save hotel data to Excel file using openpyxl"""
    try:
        # Create file if it doesn't exist
        create_hotel_excel_file_if_not_exists()
        
        # Load existing workbook
        wb = load_workbook(HOTEL_EXCEL_FILE_PATH)
        ws = wb[HOTEL_SHEET_NAME]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Prepare data for the new row
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [
            data.get('hotel_code', ''),
            data.get('name', ''),
            data.get('rating', 0),
            data.get('address', ''),
            data.get('city_id', ''),
            data.get('country_code', ''),
            data.get('map_lat', 0),
            data.get('map_lon', 0),
            json.dumps(data.get('facilities', {})),
            json.dumps(data.get('images', [])),
            timestamp
        ]
        
        # Add data to the worksheet
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        wb.save(HOTEL_EXCEL_FILE_PATH)
        print(f"Hotel data saved to Excel: {HOTEL_EXCEL_FILE_PATH}")
        return True
        
    except Exception as e:
        print(f"Error saving hotel data to Excel: {str(e)}")
        return False

def create_room_excel_file_if_not_exists():
    """Create room Excel file with headers if it doesn't exist"""
    if not os.path.exists(ROOM_EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = ROOM_SHEET_NAME
        
        # Add headers
        for col, header in enumerate(ROOM_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Save the workbook
        wb.save(ROOM_EXCEL_FILE_PATH)
        print(f"Created new room Excel file: {ROOM_EXCEL_FILE_PATH}")

def save_room_to_excel(data):
    """Save room data to Excel file using openpyxl"""
    try:
        # Create file if it doesn't exist
        create_room_excel_file_if_not_exists()
        
        # Load existing workbook
        wb = load_workbook(ROOM_EXCEL_FILE_PATH)
        ws = wb[ROOM_SHEET_NAME]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Prepare data for the new row
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [
            data.get('room_id', ''),
            data.get('hotel_code', ''),
            data.get('booking_code', ''),
            data.get('room_name', ''),
            data.get('base_price', 0),
            data.get('total_fare', 0),
            data.get('currency', ''),
            data.get('is_refundable', False),
            json.dumps(data.get('day_rates', {})),
            json.dumps(data.get('extras', {})),
            timestamp
        ]
        
        # Add data to the worksheet
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        wb.save(ROOM_EXCEL_FILE_PATH)
        print(f"Room data saved to Excel: {ROOM_EXCEL_FILE_PATH}")
        return True
        
    except Exception as e:
        print(f"Error saving room data to Excel: {str(e)}")
        return False

def create_wishlist_excel_file_if_not_exists():
    """Create wishlist Excel file with headers if it doesn't exist"""
    if not os.path.exists(WISHLIST_EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = WISHLIST_SHEET_NAME
        
        # Add headers
        for col, header in enumerate(WISHLIST_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Save the workbook
        wb.save(WISHLIST_EXCEL_FILE_PATH)
        print(f"Created new wishlist Excel file: {WISHLIST_EXCEL_FILE_PATH}")

def save_wishlist_to_excel(data):
    """Save wishlist item to Excel file using openpyxl"""
    try:
        # Create file if it doesn't exist
        create_wishlist_excel_file_if_not_exists()
        
        # Load existing workbook
        wb = load_workbook(WISHLIST_EXCEL_FILE_PATH)
        ws = wb[WISHLIST_SHEET_NAME]
        
        # Check if hotel already exists in wishlist for this customer
        for row in range(2, ws.max_row + 1):
            customer_id = ws.cell(row=row, column=2).value
            hotel_code = ws.cell(row=row, column=3).value
            if str(customer_id) == str(data.get('customer_id')) and str(hotel_code) == str(data.get('hotel_code')):
                print(f"Hotel {hotel_code} already in wishlist for customer {customer_id}")
                return True
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Generate wishlist ID
        wishlist_id = f"WL{next_row - 1:05d}"
        
        # Prepare data for the new row
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [
            wishlist_id,
            data.get('customer_id', ''),
            data.get('hotel_code', ''),
            data.get('hotel_name', ''),
            data.get('hotel_rating', 0),
            data.get('address', ''),
            data.get('city', ''),
            data.get('country', ''),
            data.get('price', 0),
            data.get('currency', 'USD'),
            data.get('image_url', ''),
            json.dumps(data.get('search_params', {})),
            timestamp
        ]
        
        # Add data to the worksheet
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        wb.save(WISHLIST_EXCEL_FILE_PATH)
        print(f"Wishlist item saved to Excel: {WISHLIST_EXCEL_FILE_PATH}")
        return True
        
    except Exception as e:
        print(f"Error saving wishlist item to Excel: {str(e)}")
        return False

def get_wishlist_by_customer(customer_id):
    """Get all wishlist items for a customer"""
    try:
        if not os.path.exists(WISHLIST_EXCEL_FILE_PATH):
            return []
        
        df = pd.read_excel(WISHLIST_EXCEL_FILE_PATH, sheet_name=WISHLIST_SHEET_NAME)
        
        # Filter by customer_id
        customer_wishlist = df[df['Customer ID'].astype(str) == str(customer_id)]
        
        return customer_wishlist.to_dict('records')
        
    except Exception as e:
        print(f"Error retrieving wishlist: {str(e)}")
        return []

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "success",
        "message": "Hotel booking backend is running"
    }), 200

@app.route('/hotel/add-hotel', methods=['POST'])
def add_hotel():
    """Handle hotel data submission and store in Excel"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "message": "No data provided"
            }), 400
        
        # Validate required fields
        required_fields = ['hotel_code', 'name', 'rating', 'address']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                "success": False,
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Save hotel data to Excel
        if save_hotel_to_excel(data):
            return jsonify({
                "success": True,
                "message": "Hotel added successfully"
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": "Failed to save hotel data"
            }), 500
            
    except Exception as e:
        print(f"Error in add_hotel endpoint: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Server error while adding hotel"
        }), 500

@app.route('/hotelRoom/add', methods=['POST'])
def add_room():
    """Handle room data submission and store in Excel"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "message": "No data provided"
            }), 400
        
        # Validate required fields
        required_fields = ['room_id', 'hotel_code', 'booking_code', 'room_name']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                "success": False,
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Save room data to Excel
        if save_room_to_excel(data):
            return jsonify({
                "success": True,
                "message": "Hotel room added successfully"
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": "Failed to save room data"
            }), 500
            
    except Exception as e:
        print(f"Error in add_room endpoint: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Server error while adding room"
        }), 500

@app.route('/hotels', methods=['GET'])
def get_hotels():
    """Get all hotels"""
    try:
        if not os.path.exists(HOTEL_EXCEL_FILE_PATH):
            return jsonify({
                "success": True,
                "data": [],
                "message": "No hotels found"
            }), 200
        
        df = pd.read_excel(HOTEL_EXCEL_FILE_PATH, sheet_name=HOTEL_SHEET_NAME)
        hotels = df.to_dict('records')
        
        return jsonify({
            "success": True,
            "data": hotels,
            "count": len(hotels)
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error retrieving hotels: {str(e)}"
        }), 500

@app.route('/rooms', methods=['GET'])
def get_rooms():
    """Get all rooms"""
    try:
        if not os.path.exists(ROOM_EXCEL_FILE_PATH):
            return jsonify({
                "success": True,
                "data": [],
                "message": "No rooms found"
            }), 200
        
        df = pd.read_excel(ROOM_EXCEL_FILE_PATH, sheet_name=ROOM_SHEET_NAME)
        rooms = df.to_dict('records')
        
        return jsonify({
            "success": True,
            "data": rooms,
            "count": len(rooms)
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error retrieving rooms: {str(e)}"
        }), 500

@app.route('/wishlist/add', methods=['POST'])
def add_to_wishlist():
    """Add hotel to user's wishlist"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "message": "No data provided"
            }), 400
        
        # Validate required fields
        required_fields = ['customer_id', 'hotel_code']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                "success": False,
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Save wishlist item to Excel
        if save_wishlist_to_excel(data):
            return jsonify({
                "success": True,
                "message": "Hotel added to wishlist successfully"
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": "Failed to add hotel to wishlist"
            }), 500
            
    except Exception as e:
        print(f"Error in add_to_wishlist endpoint: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Server error while adding to wishlist"
        }), 500

@app.route('/wishlist/<customer_id>', methods=['GET'])
def get_wishlist(customer_id):
    """Get wishlist for a specific customer"""
    try:
        wishlist_items = get_wishlist_by_customer(customer_id)
        
        return jsonify({
            "success": True,
            "data": wishlist_items,
            "count": len(wishlist_items)
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error retrieving wishlist: {str(e)}"
        }), 500

@app.route('/wishlist/remove', methods=['DELETE', 'POST'])
def remove_from_wishlist():
    """Remove hotel from user's wishlist"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "message": "No data provided"
            }), 400
        
        # Validate required fields
        required_fields = ['customer_id', 'hotel_code']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                "success": False,
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Remove from Excel
        if not os.path.exists(WISHLIST_EXCEL_FILE_PATH):
            return jsonify({
                "success": False,
                "message": "Wishlist file not found"
            }), 404
        
        wb = load_workbook(WISHLIST_EXCEL_FILE_PATH)
        ws = wb[WISHLIST_SHEET_NAME]
        
        # Find and delete the row
        row_deleted = False
        for row in range(2, ws.max_row + 1):
            customer_id = ws.cell(row=row, column=2).value
            hotel_code = ws.cell(row=row, column=3).value
            if str(customer_id) == str(data.get('customer_id')) and str(hotel_code) == str(data.get('hotel_code')):
                ws.delete_rows(row, 1)
                row_deleted = True
                break
        
        if row_deleted:
            wb.save(WISHLIST_EXCEL_FILE_PATH)
            print(f"Removed hotel {data.get('hotel_code')} from wishlist for customer {data.get('customer_id')}")
            return jsonify({
                "success": True,
                "message": "Hotel removed from wishlist successfully"
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": "Hotel not found in wishlist"
            }), 404
            
    except Exception as e:
        print(f"Error in remove_from_wishlist endpoint: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Server error while removing from wishlist"
        }), 500

def init_app():
    """Initialize application - called on startup (for Gunicorn)"""
    create_hotel_excel_file_if_not_exists()
    create_room_excel_file_if_not_exists()
    create_wishlist_excel_file_if_not_exists()
    
    print("✅ Hotel Booking Backend Server Initialized")
    print(f"📁 Hotel Excel: {os.path.abspath(HOTEL_EXCEL_FILE_PATH)}")
    print(f"📁 Room Excel: {os.path.abspath(ROOM_EXCEL_FILE_PATH)}")
    print(f"📁 Wishlist Excel: {os.path.abspath(WISHLIST_EXCEL_FILE_PATH)}")
    print("\n🌐 Available API Endpoints:")
    print("  POST /hotel/add-hotel")
    print("  POST /hotelRoom/add")
    print("  GET /hotels")
    print("  GET /rooms")
    print("  POST /wishlist/add")
    print("  POST /wishlist/remove")
    print("  GET /wishlist/<customer_id>")
    print("  GET /health")
    print("  POST /api/telr/create-order")
    print("  POST /api/telr/check-status")
    print("  POST /api/telr/webhook")

# Initialize on import (Gunicorn will call this)
init_app()

if __name__ == '__main__':
    # Development mode only - use Gunicorn for production
    print("\n⚠️  DEVELOPMENT MODE - Not for production!")
    print("For production, use: gunicorn -c gunicorn.conf.py app:app\n")
    app.run(debug=True, host='127.0.0.1', port=5001)
